"""
Detector de Drift de Fonte — Análise Preditiva Pré-Atualização.

Compara o schema esperado (definido no Power Query M do .pbit) com o schema
real da fonte de dados (Excel local) para detectar mudanças ANTES de atualizar
o dashboard no Power BI.

Suporta fontes:
- Excel (.xlsx, .xls) via File.Contents() no M
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Optional

import structlog

logger = structlog.get_logger(__name__)


# ── Enums e modelos ──────────────────────────────────────────

class DriftType(str, Enum):
    """Tipo de mudança detectada na fonte."""
    COLUMN_REMOVED = "coluna_removida"
    COLUMN_ADDED = "coluna_adicionada"
    COLUMN_RENAMED = "coluna_possivelmente_renomeada"
    TYPE_CHANGED = "tipo_alterado"
    SOURCE_NOT_FOUND = "fonte_nao_encontrada"


class DriftSeverity(str, Enum):
    """Severidade do drift."""
    CRITICAL = "critico"      # Vai quebrar o dashboard
    WARNING = "aviso"         # Pode causar problemas
    INFO = "informativo"      # Mudança sem impacto imediato


@dataclass
class ExpectedColumn:
    """Coluna esperada pelo Power Query M."""
    name: str
    m_type: str  # Tipo no M (type text, type number, etc.)
    python_type: str = ""  # Mapeamento para tipo Python/pandas


@dataclass
class SourceInfo:
    """Informações sobre uma fonte de dados extraída do M."""
    table_name: str
    source_type: str  # "excel", "csv", etc.
    file_path: str
    sheet_name: Optional[str] = None
    expected_columns: list[ExpectedColumn] = field(default_factory=list)
    m_expression: str = ""


@dataclass
class DriftItem:
    """Um item de drift detectado."""
    table_name: str
    drift_type: DriftType
    severity: DriftSeverity
    column_name: str
    detail: str
    expected_type: Optional[str] = None
    actual_type: Optional[str] = None
    suggestion: Optional[str] = None


@dataclass
class DriftReport:
    """Relatório completo de drift."""
    sources_analyzed: int = 0
    sources_ok: int = 0
    sources_with_drift: int = 0
    sources_not_found: int = 0
    drifts: list[DriftItem] = field(default_factory=list)
    source_details: list[SourceInfo] = field(default_factory=list)

    @property
    def has_critical(self) -> bool:
        return any(d.severity == DriftSeverity.CRITICAL for d in self.drifts)

    @property
    def critical_count(self) -> int:
        return sum(1 for d in self.drifts if d.severity == DriftSeverity.CRITICAL)

    @property
    def warning_count(self) -> int:
        return sum(1 for d in self.drifts if d.severity == DriftSeverity.WARNING)

    @property
    def info_count(self) -> int:
        return sum(1 for d in self.drifts if d.severity == DriftSeverity.INFO)


# ── Mapeamento de tipos M → pandas ──────────────────────────

M_TYPE_MAP = {
    "type text": "object",
    "type number": "float64",
    "type date": "datetime64",
    "type datetime": "datetime64",
    "type logical": "bool",
    "type any": "object",
    "Int64.Type": "int64",
    "Int32.Type": "int32",
    "Percentage.Type": "float64",
    "Currency.Type": "float64",
}


# ── Parser de Power Query M ─────────────────────────────────

def parse_m_sources(m_expression: str, table_name: str) -> Optional[SourceInfo]:
    """
    Extrai informações de fonte de um Power Query M.

    Detecta padrões:
    - Excel.Workbook(File.Contents("caminho"), ...)
    - Table.TransformColumnTypes(... {{"col", type}})
    """
    if not m_expression:
        return None

    # 1. Detectar fonte Excel
    file_match = re.search(
        r'File\.Contents\(\s*"([^"]+)"\s*\)',
        m_expression,
    )
    if not file_match:
        return None  # Fonte não é arquivo local

    file_path = file_match.group(1)

    # Verificar se é Excel
    ext = Path(file_path).suffix.lower()
    if ext not in (".xlsx", ".xls", ".xlsm", ".xlsb"):
        return None

    # 2. Detectar sheet name
    sheet_name = None
    sheet_match = re.search(
        r'\{[^}]*Item\s*=\s*"([^"]+)"\s*,\s*Kind\s*=\s*"Sheet"',
        m_expression,
    )
    if sheet_match:
        sheet_name = sheet_match.group(1)

    # 3. Extrair colunas esperadas do Table.TransformColumnTypes
    expected_columns = _parse_column_types(m_expression)

    source = SourceInfo(
        table_name=table_name,
        source_type="excel",
        file_path=file_path,
        sheet_name=sheet_name,
        expected_columns=expected_columns,
        m_expression=m_expression,
    )

    logger.info(
        "Fonte Excel detectada",
        table=table_name,
        path=file_path,
        sheet=sheet_name,
        columns=len(expected_columns),
    )

    return source


def _parse_column_types(m_expression: str) -> list[ExpectedColumn]:
    """
    Extrai colunas e tipos do Table.TransformColumnTypes.

    Padrão M: {{"NomeColuna", type text}, {"Outra", type number}}
    """
    columns = []
    seen = set()

    # Encontra TODOS os blocos Table.TransformColumnTypes
    pattern = r'Table\.TransformColumnTypes\s*\([^,]+,\s*\{((?:\{[^}]+\}\s*,?\s*)+)\}'
    matches = re.finditer(pattern, m_expression, re.DOTALL)

    for match in matches:
        block = match.group(1)
        # Extrair cada par {"nome", tipo}
        col_pattern = r'\{\s*"([^"]+)"\s*,\s*([\w.]+(?:\s+\w+)?)\s*\}'
        for col_match in re.finditer(col_pattern, block):
            col_name = col_match.group(1)
            col_type = col_match.group(2).strip()

            if col_name not in seen:
                seen.add(col_name)
                columns.append(ExpectedColumn(
                    name=col_name,
                    m_type=col_type,
                    python_type=M_TYPE_MAP.get(col_type, "object"),
                ))

    return columns


# ── Detector de Drift ────────────────────────────────────────

def detect_drift(source: SourceInfo) -> list[DriftItem]:
    """
    Compara o schema esperado com o schema real do Excel.

    Returns:
        Lista de drifts detectados.
    """
    drifts: list[DriftItem] = []

    file_path = Path(source.file_path)
    if not file_path.exists():
        drifts.append(DriftItem(
            table_name=source.table_name,
            drift_type=DriftType.SOURCE_NOT_FOUND,
            severity=DriftSeverity.CRITICAL,
            column_name="*",
            detail=f"Arquivo não encontrado: {source.file_path}",
            suggestion="Verifique se o arquivo Excel existe no caminho especificado.",
        ))
        return drifts

    # Ler headers reais do Excel
    try:
        import openpyxl
        actual_columns = _read_excel_headers(file_path, source.sheet_name)
    except ImportError:
        try:
            import pandas as pd
            actual_columns = _read_excel_headers_pandas(file_path, source.sheet_name)
        except Exception as e:
            logger.error("Erro ao ler Excel", error=str(e))
            return drifts
    except Exception as e:
        logger.error("Erro ao ler Excel", error=str(e))
        drifts.append(DriftItem(
            table_name=source.table_name,
            drift_type=DriftType.SOURCE_NOT_FOUND,
            severity=DriftSeverity.CRITICAL,
            column_name="*",
            detail=f"Erro ao ler arquivo: {e}",
        ))
        return drifts

    expected_names = {c.name for c in source.expected_columns}
    actual_names = set(actual_columns)

    # Colunas removidas (CRÍTICO — vai quebrar)
    removed = expected_names - actual_names
    for col in sorted(removed):
        # Tentar detectar renomeação (fuzzy match)
        added_only = actual_names - expected_names
        possible_rename = _find_similar(col, added_only)

        if possible_rename:
            drifts.append(DriftItem(
                table_name=source.table_name,
                drift_type=DriftType.COLUMN_RENAMED,
                severity=DriftSeverity.CRITICAL,
                column_name=col,
                detail=f"Coluna '{col}' não encontrada. Possível renomeação para '{possible_rename}'.",
                suggestion=f"Renomear referências de '{col}' para '{possible_rename}' no modelo.",
            ))
        else:
            drifts.append(DriftItem(
                table_name=source.table_name,
                drift_type=DriftType.COLUMN_REMOVED,
                severity=DriftSeverity.CRITICAL,
                column_name=col,
                detail=f"Coluna '{col}' esperada pelo modelo mas não existe mais na fonte.",
                suggestion="Remova referências a esta coluna ou atualize a fonte.",
            ))

    # Colunas adicionadas (INFO — não quebra, mas pode indicar mudança)
    added = actual_names - expected_names
    # Filtra colunas que já foram marcadas como possível renomeação
    renamed_targets = set()
    for d in drifts:
        if d.drift_type == DriftType.COLUMN_RENAMED and d.suggestion:
            target = d.suggestion.split("'")
            if len(target) >= 4:
                renamed_targets.add(target[3])

    for col in sorted(added - renamed_targets):
        drifts.append(DriftItem(
            table_name=source.table_name,
            drift_type=DriftType.COLUMN_ADDED,
            severity=DriftSeverity.INFO,
            column_name=col,
            detail=f"Nova coluna '{col}' encontrada na fonte (não existe no modelo).",
            suggestion="Considere adicionar esta coluna ao modelo se necessário.",
        ))

    logger.info(
        "Drift analysis complete",
        table=source.table_name,
        expected=len(expected_names),
        actual=len(actual_names),
        removed=len(removed),
        added=len(added),
    )

    return drifts


def _read_excel_headers(file_path: Path, sheet_name: Optional[str] = None) -> list[str]:
    """Lê headers do Excel usando openpyxl (sem carregar dados)."""
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Ler as primeiras linhas para detectar headers
        # Power Query com PromoteHeaders pula linhas vazias
        headers = []
        for row in ws.iter_rows(max_row=5, values_only=True):
            candidate = [str(c).strip() if c is not None else "" for c in row]
            # Um bom header tem maioria de strings não-vazias
            non_empty = sum(1 for c in candidate if c and c != "None")
            if non_empty > len(candidate) * 0.5:
                headers = [c for c in candidate if c and c != "None"]
                break

        return headers
    finally:
        wb.close()


def _read_excel_headers_pandas(file_path: Path, sheet_name: Optional[str] = None) -> list[str]:
    """Fallback: lê headers usando pandas."""
    import pandas as pd

    kwargs = {"nrows": 0}
    if sheet_name:
        kwargs["sheet_name"] = sheet_name

    df = pd.read_excel(str(file_path), **kwargs)
    return list(df.columns)


def _find_similar(name: str, candidates: set[str], threshold: float = 0.6) -> Optional[str]:
    """
    Busca um nome similar entre os candidatos (detecção de renomeação).
    Usa similaridade de sequência simples.
    """
    if not candidates:
        return None

    name_lower = name.lower().replace(" ", "").replace("_", "")
    best_match = None
    best_score = 0.0

    for candidate in candidates:
        cand_lower = candidate.lower().replace(" ", "").replace("_", "")

        # Substring match
        if name_lower in cand_lower or cand_lower in name_lower:
            score = min(len(name_lower), len(cand_lower)) / max(len(name_lower), len(cand_lower))
            if score > best_score:
                best_score = score
                best_match = candidate
                continue

        # Character overlap
        common = set(name_lower) & set(cand_lower)
        total = set(name_lower) | set(cand_lower)
        if total:
            score = len(common) / len(total)
            if score > best_score:
                best_score = score
                best_match = candidate

    if best_score >= threshold:
        return best_match
    return None


# ── Função principal ─────────────────────────────────────────

def analyze_source_drift(schema_tables: list[dict]) -> DriftReport:
    """
    Analisa drift de fontes para todas as tabelas do schema.

    Args:
        schema_tables: Lista de tabelas do schema (do .pbit DataModelSchema).
                       Cada tabela deve ter 'name' e 'partitions' com 'source.expression'.

    Returns:
        DriftReport com todos os drifts encontrados.
    """
    report = DriftReport()

    for table in schema_tables:
        table_name = table.get("name", "?")

        # Pular tabelas técnicas
        if table_name.startswith("LocalDateTable_") or table_name.startswith("DateTableTemplate"):
            continue

        # Extrair M expression
        for partition in table.get("partitions", []):
            source = partition.get("source", {})
            expr = source.get("expression", "")
            if isinstance(expr, list):
                expr = "\n".join(expr)
            if not expr:
                continue

            source_info = parse_m_sources(expr, table_name)
            if not source_info:
                continue

            report.sources_analyzed += 1
            report.source_details.append(source_info)

            drifts = detect_drift(source_info)
            report.drifts.extend(drifts)

            if any(d.drift_type == DriftType.SOURCE_NOT_FOUND for d in drifts):
                report.sources_not_found += 1
            elif drifts:
                report.sources_with_drift += 1
            else:
                report.sources_ok += 1

    return report
